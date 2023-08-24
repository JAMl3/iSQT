import subprocess
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

import openpyxl
import pandas as pd
import pyodbc
from PIL import Image, ImageTk
from ttkthemes import ThemedTk

conn = None
cursor = None


class CredentialsDialog(tk.simpledialog.Dialog):
    def body(self, master):
        ttk.Label(master, text="Enter your Aurora username and password").pack()

        self.username_entry = ttk.Entry(master)
        self.password_entry = ttk.Entry(master, show="*")

        self.username_entry.pack()
        self.password_entry.pack()

        return self.username_entry

    def apply(self):
        self.username = self.username_entry.get()
        self.password = self.password_entry.get()


def execute_query():
    selected_table = query_combobox.get()
    clear_search_status_label()
    if not selected_table:
        messagebox.showerror("Error", "Please select a table.")
        return
    search_entry.delete(0, tk.END)
    try:
        selected_columns = column_listbox.get(0, tk.END)
        column_list_str = ", ".join(selected_columns)

        query = f"SELECT {column_list_str} FROM {selected_table}"

        cursor.execute(query)
        data = cursor.fetchall()

        result_text.delete(1.0, tk.END)
        if data:
            columns = [column[0] for column in cursor.description]
            result_text.insert(tk.END, "\t".join(columns) + "\n")
            for row in data:
                result_text.insert(tk.END, "\t".join(map(str, row)) + "\n")
        else:
            result_text.insert(tk.END, "No data available.")

    except pyodbc.Error as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

    search_status_label.config(text="")


def export_to_excel():
    selected_table = query_combobox.get()
    if not selected_table:
        messagebox.showerror("Error", "Please select a query.")
        return

    selected_columns = column_listbox.get(0, tk.END)
    column_list_str = ", ".join(selected_columns)

    try:
        query = f"SELECT {column_list_str} FROM {selected_table}"
        data_frame = pd.read_sql(query, conn)

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialdir="/path/to/your/directory",
                                                 initialfile=f"{selected_table}.xlsx")
        if file_path:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            for col_idx, column in enumerate(selected_columns, start=1):
                worksheet.cell(row=1, column=col_idx, value=column)

            for row_idx, row in enumerate(data_frame.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

            workbook.save(file_path)
            workbook.close()

            subprocess.Popen(["start", "excel", file_path], shell=True)

    except pyodbc.Error as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def move_column(direction):
    selected_index = column_listbox.curselection()
    if not selected_index:
        return

    selected_index = selected_index[0]
    selected_column = column_listbox.get(selected_index)
    column_listbox.delete(selected_index)

    new_index = selected_index + direction
    if new_index < 0:
        new_index = 0
    elif new_index > column_listbox.size():
        new_index = column_listbox.size()

    column_listbox.insert(new_index, selected_column)
    column_listbox.selection_set(new_index)


def update_available_columns(event):
    selected_table = query_combobox.get()
    if selected_table:
        column_listbox.delete(0, tk.END)
        columns = [column.column_name for column in cursor.columns(table=selected_table)]
        for column in columns:
            column_listbox.insert(tk.END, column)


def get_credentials_and_connect():
    global conn, cursor
    credentials_dialog = CredentialsDialog(root)
    user_input = credentials_dialog.username
    password_input = credentials_dialog.password

    if user_input and password_input:
        try:
            conn = pyodbc.connect(f'DRIVER={"iSeries Access ODBC Driver"};SYSTEM=YOUR_SERVER_IP;'
                                  f'DBQ=YOUR_LIBRARY_NAME;DFTPKGLIB=QGPL;LANGUAGEID=ENU;PKG=QGPL/DEFAULT(IBM),2,0,1,0,512;'
                                  f'UID={user_input};PWD={password_input}')
            cursor = conn.cursor()

            connected_label.config(text=f"{user_input} Connected!".upper(), foreground="green")

            table_list = [table.table_name for table in cursor.tables()]
            query_combobox['values'] = table_list

            query_combobox.bind("<<ComboboxSelected>>", update_available_columns)

            if table_list:
                first_table_columns = [column.column_name for column in cursor.columns(table=table_list[0])]
                column_listbox.delete(0, tk.END)
                for column in first_table_columns:
                    column_listbox.insert(tk.END, column)

        except pyodbc.Error as e:
            messagebox.showerror("Error", f"Could not connect to the database: {e}")


def search_results():
    global search_indices, current_search_index

    search_term = search_entry.get()
    if search_term:
        result_text.tag_remove("highlight", "1.0", tk.END)
        search_indices = []

        start_index = "1.0"
        while True:
            start_index = result_text.search(search_term, start_index, stopindex=tk.END, nocase=True)
            if not start_index:
                break
            end_index = f"{start_index}+{len(search_term)}c"
            search_indices.append((start_index, end_index))
            start_index = end_index

        if search_indices:
            current_search_index = 0
            start, _ = search_indices[current_search_index]
            result_text.tag_add("highlight", start, f"{start}+{len(search_term)}c")
            result_text.tag_config("highlight", background="yellow", foreground="black")
            result_text.see(start)
            search_status_label.config(text="")

            if len(search_indices) > 1:
                search_status_label.config(text=f"Found {len(search_indices)} matches")

        else:
            search_status_label.config(text="No Results Found!")


def search_next():
    global search_indices, current_search_index

    if search_indices and len(search_indices) > 1:
        current_search_index = (current_search_index + 1) % len(search_indices)
        start, end = search_indices[current_search_index]
        result_text.tag_remove("highlight", "1.0", tk.END)
        result_text.tag_add("highlight", start, end)
        result_text.tag_config("highlight", background="yellow", foreground="black")
        result_text.see(start)


def copy_to_clipboard():
    result_text_contents = result_text.get("1.0", tk.END)
    root.clipboard_clear()
    root.clipboard_append(result_text_contents)
    root.update()  # To update the clipboard content

    search_status_label.config(text="Results Copied")
    root.after(2000, clear_search_status_label)  # Clear the message after 2000 milliseconds (2 seconds)


def clear_search_status_label():
    search_status_label.config(text="")


current_search_index = 0  # Global variable to keep track of the current search index

root = ThemedTk(theme="arc")
root.title("iSeries Query Tool")
root.resizable(False, False)

# Load and display the image logo
logo_image = Image.open("company_Logo.png")
resized_logo = logo_image.resize((200, 100))  # Adjust the size as needed
logo_photo = ImageTk.PhotoImage(resized_logo)


main_frame = ttk.Frame(root)
main_frame.pack(padx=2, pady=2)

left_frame = ttk.Frame(main_frame)
left_frame.pack(side=tk.LEFT, padx=2, pady=2)

logo_label = ttk.Label(left_frame, image=logo_photo)
logo_label.pack()

connected_label = ttk.Label(left_frame, text="", foreground="green", font=12)
connected_label.pack()

connect_button = ttk.Button(left_frame, text="Connect", command=get_credentials_and_connect)
connect_button.pack()

query_label = ttk.Label(left_frame, text="Select Query:")
query_label.pack()

query_combobox = ttk.Combobox(left_frame)
query_combobox.pack()

column_label = ttk.Label(left_frame, text="Column Layout:")
column_label.pack()

scrollbar_y1 = ttk.Scrollbar(left_frame)
scrollbar_y1.pack(side=tk.RIGHT, fill=tk.Y)

column_listbox = tk.Listbox(left_frame, yscrollcommand=scrollbar_y1.set, selectmode=tk.MULTIPLE)
column_listbox.pack(fill=tk.BOTH, expand=True)
scrollbar_y1.config(command=column_listbox.yview)

move_buttons_frame = ttk.Frame(left_frame)
move_up_button = ttk.Button(move_buttons_frame, text="Move Up", command=lambda: move_column(-1))
move_up_button.pack(side=tk.LEFT)

move_down_button = ttk.Button(move_buttons_frame, text="Move Down", command=lambda: move_column(1))
move_down_button.pack(side=tk.LEFT)

move_buttons_frame.pack()

execute_button = ttk.Button(left_frame, text="Show Results", command=execute_query)
execute_button.pack()

right_frame = ttk.Frame(main_frame)
right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

scrollbar_y = ttk.Scrollbar(right_frame)
scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

scrollbar_x = ttk.Scrollbar(right_frame, orient=tk.HORIZONTAL)
scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)

result_text = tk.Text(right_frame, wrap=tk.NONE, yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
result_text.pack(fill=tk.BOTH, expand=True)

scrollbar_y.config(command=result_text.yview)
scrollbar_x.config(command=result_text.xview)

search_entry = ttk.Entry(right_frame)
search_entry.pack(side=tk.LEFT)

search_button = ttk.Button(right_frame, text="Search", command=search_results)
search_button.pack(side=tk.LEFT)

search_next_button = ttk.Button(right_frame, text="Next", command=search_next)
search_next_button.pack(side=tk.LEFT)

search_status_label = ttk.Label(right_frame, text="", foreground="red")
search_status_label.pack(side=tk.LEFT)

export_button = ttk.Button(right_frame, text="Export To Excel", command=export_to_excel)
export_button.pack(side=tk.RIGHT)

copy_to_clipboard_button = ttk.Button(right_frame, text="Copy to Clipboard", command=copy_to_clipboard)
copy_to_clipboard_button.pack(side=tk.RIGHT)

root.mainloop()
